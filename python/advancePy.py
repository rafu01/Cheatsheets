# importing another py file
import openpyxl
from selenium import webdriver
from bs4 import BeautifulSoup
import requests
from email.mime.image import MIMEImage
import smtplib
from email.mime.text import MIMEText
# mime = multipurpose internet mail extension
from email.mime.multipart import MIMEMultipart
import webbrowser
import string
import random
from datetime import datetime
import json
import csv
import shutil
from pathlib import Path
from sales import func1, func2
# or
import sales
# now use like this
sales.func1()
# python standard library
path = Path("hello.py")
path.exists()  # returns boolean if this directory exists or not
path.is_file()  # to check if this path represents a file or not
path.is_dir()  # to check if this path represents a directory or no
path.absolute()  # returns the absolute directory of the path
path.name  # returns the file name
path.parent  # returns the parent folder
path.suffix  # returns the extension
path.stem  # returns the file name w/o the extension
path.with_name("file.txt")  # creates a file
path.rename()  # pass parameters of the new name
path.read_text()  # returns the content of the file as string
path.write_text()  # writes file
#  for copying a file use shutil module
shutil.copy(source, target)

# writing csv file
with open("daraz.csv", "w") as file:  # opening a csv, "w" means to write the file
    writer = csv.writer(file)
    writer.writerow("a", "b")  # 1 row 2 colums will be added
# to read a csv
with open("daraz.csv", "w") as file:
    reader = csv.reader(file)
    # will convert all data into  a list of lists (each line is a list)
    print(list(reader))
    for row in reader:  # reading by row
        print(row)
# working with json files
movies = [{"id": 1, "name": "a"}, {
    "id": 2, "name": "b"}, {"id": 3, "name": "c"}]
movie_file = json.dumps(movies)  # converts it to json data
# create a json file
Path("movies.json").write_text(movie_file)
# read a json file
readingJsonFile = Path("movies.json").read_text()
# converts it to list of dictionaries
movies_dictionaries = json.loads(readingJsonFile)
# data time
now = datetime.now()
# generate random valus
random.random()  # generates a double
random.randint(1, 10)  # generates between this range
random.choice([1, 25, 7])  # picks a number from this list
random.choices([1, 5, 3, 9, 6], k=2)  # pick k numbers from this list
# generating password using random
# join metho joins these strings
"".join(random.choices(["abaaabdkdojiwjiojw"], k=6))
string.ascii_letters  # this returns all ascii characters
string.digits  # returns 0 to 9
"".join(random.choices(string.ascii_letters+string.digits, k=8))
num = [1, 2, 3, 4]
# shuffle a list
random.shuffle(num)
# open a browser
webbrowser.open("http://google.com")
# sending emails
message = MIMEMultipart()
message["from"] = "Nayeem Rafsan"
message["to"] = "teset@gamil.com"
message["subject"] = "this is testing"
message.attach(MIMEText("Body"))  # can't add body of the message
# we have attach the txt, image file here
# we have created a text with MIMEText constructor
# now we need a SMTP server

with smtplib.SMTP(host="smtp.gmail.com", port=465) as smtp:
    smtp.ehlo()  # sending hello message to server
    smtp.starttls()  # puts smtp to transport layer security
    smtp.login("username@gmail.com", "passwordhere")
    smtp.send_message(message)
    print("sent")  # for confirmation
# attach image
# this only takes byte so convert image to byte
message.attach(MIMEImage(Path("image1.jpg").read_bytes))
# first https://myaccount.google.com/u/1/lesssecureapps and turn off less secure app

# virtual environments:
# for certain dependencies (ex: requests is version 20.0 bt you need 18.0 for your app)
# so we create a virtual environment and create these dependencies
# for this we use pipenv
# pip3 install pipenv
# then activate it

# YELP API
# install requests using pip3 or pipenv

url = "https://api.yelp.com/v3/businesses/search"
api_key = "aaaaaaaaaaaaaaaaaaaaa"
head = {
    "Authorization": "Bearer "+api_key
}
param = {
    "location": "NYC"
}
# this will take 3 parameters to work, url, header, params
# this will return a json file
r = requests.get(url, headers=head, params=param)
# now extract information from the json file
r.json()  # this converts the json to dictionary
business = r.json()["businesses"]
for b in business:
    print(b['name'])
# or use list comphrehension
[b["name"] for b in business]
[b["name"] for b in business if business["rating"] > 4]  # filtering the businesses
# hiding the API key
# create config.py create variable for api_key
# import config in app.py and replace api_key with config.api_key
# create .gitignore file and write config.py
# webscarping

req = requests.get("http://stackoverflow.com/questions")

soup = BeautifulSoup(req.text, "html.parser")
# find container thqat contains all info
qs = soup.select(".question-summary")  # . dot means class
# under question-summary
print(qs[0].select_one(".question-hyperlink").getText())
for q in qs:
    print(q.select_one(".question-hyperlink").getText())
# browser automation using Selenium

browser = webdriver.Chrome()
browser.get("http://github.com")

# we can get elements by their className/ tags or text
# get elements by text
sign_in = browser.find_elements_by_link_text("Sign in")
# now click it
sign_in.click()
# now get elements by id
user_name = browser.find_element_by_id("login_field")
# for typing use send_keys()
user_name.send_keys("rafu01")
user_pass = browser.find_element_by_id("password")
user_pass.send_keys("password")
user_pass.submit()

# checking if this text exists in this page or not
assert "rafu01" in browser.page_source

# working with excel sheet

wp = openpyxl.load_workbook("myExcel.xlsx")
wp.sheetnames  # prints all sheet names

sheet = wb["Sheet1"]

cell = sheet["a1"]
# or use
cell = sheet.cell(row=1, column=1)
# sheet[row,colum]
sheet[1:3]  # only row
